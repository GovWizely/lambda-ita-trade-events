import vcr
from openpyxl import load_workbook
import service


def my_matcher(a, b):
    assert True


my_vcr = vcr.VCR()
my_vcr.register_matcher('placeholder', my_matcher)


def test_ita_events():
    with my_vcr.use_cassette('test_ita_events', match_on=['placeholder']):
        entries = service.get_event_list()
        assert len(entries) == 115
        expected_entry = {
            'cost': 1995.0,
            'detaildesc': 'The Chicago MBDA Export Center is organizing a trade mission to Mozambique. They will recruit up to 12 businesses to participate in this event. MBDA is working with the U.S.Commercial service Office in Maputo, Mozambique to organise a B2B meeting, sites visits and a reception for the delegates.',  # NOQA: E501
            'eventid': '40033',
            'eventname': 'Minority Business Development Agency - Trade Mission to Mozambique',
            'eventtype': 'Certified Trade Mission',
            'evenddt': '2020-03-24',
            'registrationlink': '',
            'registrationtitle': 'Apply for this event',
            'evstartdt': '2020-03-19',
            'url': None,
            'contacts': [
                {
                    'email': 'Tamarind.Murrietta@trade.gov',
                    'firstname': 'Tamarind',
                    'lastname': 'Murrietta',
                    'title': 'Senior Commercial Officer',
                    'phone': '+258-2135-5418',
                    'post': 'Maputo'
                },
                {
                    'email': 'Daniel.Donato@trade.gov',
                    'firstname': 'Daniel',
                    'lastname': 'Donato',
                    'title': 'Commercial Specialist',
                    'phone': '',
                    'post': 'Maputo'
                },
                {
                    'email': 'Marina.Matola@trade.gov',
                    'firstname': 'Marina',
                    'lastname': 'Matola',
                    'title': 'Administrative Specialist',
                    'phone': '+258843141688',
                    'post': 'Maputo'
                }
            ],
            'venues': [
                {
                    'city': 'Maputo',
                    'country': 'Mozambique',
                    'state': '',
                    'location': 'Radisson Blu Hotel Maputo'
                }
            ],
            'industries': []
        }
        assert entries[0] == expected_entry


def test_tepp_events(monkeypatch):
    workbook = load_workbook('sample_tepp.xlsx')

    def mockreturn():
        first_sheet = workbook.sheetnames[0]
        return workbook[first_sheet]

    monkeypatch.setattr(service, "get_tepp_worksheet", mockreturn)

    entries = service.get_tepp_events()
    assert len(entries) == 32
    expected_entry = {
        'eventid': '0ddc7422238d453c18f3f643e9181788caf06257',
        'industries': ['Design & Construction'],
        'evenddt': '2020-02-07',
        'url': 'blank',
        'eventtype': 'Trade Events Partnership Program',
        'evstartdt': '2020-02-04',
        'registrationlink': None,
        'contacts': [
            {
                'firstname': 'Jackie',
                'title': None,
                'lastname': 'James',
                'phone': '972-536-6379',
                'post': 'Irving',
                'email': 'jackie.james@informa.com'
            }
        ],
        'detaildesc': None,
        'eventname': 'World of Concrete 2020',
        'venues': [
            {
                'city': 'Las Vegas',
                'state': 'NV',
                'country': 'USA',
                'location': 'Las Vegas, NV, USA'
            }
        ],
        'cost': None,
        'registrationtitle': None
    }
    assert entries[0] == expected_entry