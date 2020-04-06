# -*- coding: utf-8 -*-
import datetime as dt
import hashlib
import json
import logging
from io import BytesIO

import boto3
import openpyxl
import requests
from botocore.exceptions import ClientError
from bs4 import BeautifulSoup

TAGS = [
    "cost",
    "detaildesc",
    "eventid",
    "eventname",
    "eventtype",
    "evenddt",
    "registrationlink",
    "registrationtitle",
    "evstartdt",
]
CONTACT_TAGS = ["email", "firstname", "lastname", "title", "phone", "post"]
VENUE_TAGS = ["city", "country", "state", "location"]
ENDPOINT = "http://emenuapps.ita.doc.gov/ePublic/GetEventXML?StartDT={0}&EndDT={1}"
BUCKET = "trade-events"

s3 = boto3.client("s3")


def handler(event, context):
    entries = get_concat_events()
    try:
        s3.put_object(
            Bucket=BUCKET,
            Body=json.dumps(entries),
            Key="ita.json",
            ContentType="application/json",
        )
        print(f" ✅ Uploaded ita.json file with {len(entries):d} trade events")
    except ClientError as e:
        logging.error(e)
        return False
    return True


def get_event(item):
    event = {tag: get_text(item, tag) for tag in TAGS}
    event["url"] = get_url(item)
    event["evenddt"] = normalize_date(event["evenddt"])
    event["evstartdt"] = normalize_date(event["evstartdt"])
    event["contacts"] = get_contacts(item)
    event["venues"] = get_venues(item)
    event["industries"] = get_industries(item)
    event["cost"] = float(event["cost"])
    return event


def get_venues(item):
    return [get_venue(item.stop)]


def get_venue(venue_entry):
    venue = {tag: get_text(venue_entry, tag) for tag in VENUE_TAGS}
    return venue


def get_industries(item):
    return [get_industry(industry) for industry in item.findAll("industry")]


def get_industry(industry_entry):
    return industry_entry.text


def get_contacts(item):
    return [get_contact(contact) for contact in item.findAll("contact")]


def get_contact(contact_entry):
    contact = {tag: get_text(contact_entry, tag) for tag in CONTACT_TAGS}
    return contact


def get_url(item):
    try:
        url = item.websites.website["url"]
    except Exception:
        url = None
    return url


def normalize_date(entry_date):
    return dt.datetime.strptime(entry_date, "%m/%d/%Y").strftime("%Y-%m-%d")


def get_text(item, tag):
    inner_text = get_inner_text(item, tag)
    result_text = "{}".format(inner_text)
    return result_text


def get_inner_text(item, tag):
    element = item.find(tag)
    try:
        element_text = element.text
    except AttributeError:
        element_text = ""
    return element_text


def get_event_list():
    print("Fetching XML feed of items...")
    tomorrow = dt.date.today() + dt.timedelta(days=1)
    far_off = dt.date.today() + dt.timedelta(days=365 * 4)
    endpoint = ENDPOINT.format(
        tomorrow.strftime("%m/%d/%Y"), far_off.strftime("%m/%d/%Y")
    )
    soup = get_soup(endpoint)
    items = soup.eventlist.findAll("eventinfo")
    event_list = [get_event(item) for item in items]
    print(f"Found {len(event_list):d} ITA trade items on eMenu")
    return event_list


def get_soup(link):
    response = requests.get(link)
    poorly_formed_xml = response.text
    soup = BeautifulSoup(poorly_formed_xml, "html.parser")
    return soup


def get_tepp_worksheet():
    s3_response_object = s3.get_object(Bucket=BUCKET, Key="tepp_export.xlsx")
    response_body = s3_response_object["Body"].read()
    workbook = openpyxl.load_workbook(BytesIO(response_body))
    first_sheet = workbook.sheetnames[0]
    return workbook[first_sheet]


def get_tepp_headers():
    worksheet = get_tepp_worksheet()
    headers = {}
    idx = 0
    for COL in worksheet.iter_cols(1, worksheet.max_column):
        headers[COL[0].value] = idx
        idx += 1
    return headers


def convert_date(row, date_col):
    """
    Allow TypeError to resolve as "None", since that means the field is empty.
    ValueError should still fail, because that means the date was formatted incorrectly.
    """
    try:
        return dt.datetime.strptime(
            str(row[get_tepp_headers()[date_col]]), "%Y-%m-%d %H:%M:%S"
        ).strftime("%Y-%m-%d")
    except TypeError:
        return None


def get_first_name(row):
    try:
        return row[get_tepp_headers()["Contact Name (First and Last)"]].partition(" ")[
            0
        ]
    except AttributeError:
        return None


def get_last_name(row):
    try:
        return row[get_tepp_headers()["Contact Name (First and Last)"]].partition(" ")[
            -1
        ]
    except AttributeError:
        return None


def get_tepp_contact_info(row):
    contact = {}
    contact["firstname"] = get_first_name(row)
    contact["title"] = None
    contact["lastname"] = get_last_name(row)
    contact["phone"] = row[get_tepp_headers()["Contact Phone Number"]]
    contact["post"] = row[get_tepp_headers()["City (Organization)"]]
    contact["email"] = row[get_tepp_headers()["Contact Email"]]
    return contact


def get_tepp_venue_info(row):
    venue = {}
    venue["city"] = row[get_tepp_headers()["Event Location - City"]]
    venue["state"] = row[get_tepp_headers()["Event Location - State (U.S. Only)"]]
    venue["country"] = row[get_tepp_headers()["Event Location - Country"]]
    location_array = [venue["city"], venue["state"], venue["country"]]
    venue["location"] = ", ".join([str(item) for item in location_array if item])
    return venue


def generate_event_id(row):
    m = hashlib.sha1()
    unique_list = [
        row[get_tepp_headers()["Event Name"]],
        convert_date(row, "Event Start Date"),
        convert_date(row, "Event End Date"),
        row[get_tepp_headers()["Event Location - City"]],
    ]
    m.update("".join([str(item) for item in unique_list if item]).encode())
    return m.hexdigest()


def get_tepp_industry(row):
    if row[get_tepp_headers()["Primary Industry"]]:
        return [row[get_tepp_headers()["Primary Industry"]]]
    else:
        return []


def get_tepp_events():
    events = []
    worksheet = get_tepp_worksheet()
    permitted_status = [
        "MOA Received",
        "Waiting on End of Show Report",
        "Event Completed",
    ]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[get_tepp_headers()["Status"]] in permitted_status:
            event = {}
            event["eventid"] = generate_event_id(row)
            event["industries"] = get_tepp_industry(row)
            event["evenddt"] = convert_date(row, "Event End Date")
            event["url"] = row[get_tepp_headers()["Show Website"]]
            event["eventtype"] = "Trade Events Partnership Program"
            event["evstartdt"] = convert_date(row, "Event Start Date")
            event["registrationlink"] = None
            event["contacts"] = [get_tepp_contact_info(row)]
            event["detaildesc"] = row[get_tepp_headers()["Show Description"]]
            event["eventname"] = row[get_tepp_headers()["Event Name"]]
            event["venues"] = [get_tepp_venue_info(row)]
            event["cost"] = None
            event["registrationtitle"] = None
            events.append(event)
    print(f"Found {len(events):d} TEPP items in the spreadsheet")
    return events


def get_concat_events():
    concat_events = get_event_list()
    concat_events.extend(get_tepp_events())
    return concat_events
